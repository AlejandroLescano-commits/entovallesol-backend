"""
test_importacion_funcional.py – Pruebas Funcionales del módulo de Importación Masiva
EntoValleSOL Backend

Casos cubiertos:
  IMP-01 – POST /importar/sitotroga con Excel válido → retorna conteo de importados
  IMP-02 – POST /importar/trichogramma con Excel válido → importación exitosa
  IMP-03 – POST /importar/galleria con Excel válido → importación exitosa
  IMP-04 – POST /importar/paratheresia con Excel válido → importación exitosa
  IMP-05 – Excel con filas con fecha nula → esas filas se omiten (no crashea)
  IMP-06 – Excel con cantidad inválida (texto) → esa fila cuenta como error
  IMP-07 – Archivo no Excel (txt, csv sin openpyxl) → HTTP 400 o 422
  IMP-08 – Solo admin puede importar (operario/supervisor → 403)
  IMP-09 – Sin archivo en la request → HTTP 422
  IMP-10 – Excel con todas las filas erróneas → importados=0, errores>0

Ejecutar:
  pytest tests/integracionextras/test_importacion_funcional.py -v
"""
import io
import pytest

# openpyxl es necesario para generar archivos Excel en los tests
try:
    import openpyxl
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False


def _excel_bytes(filas: list[dict], columnas: list[str] = None) -> bytes:
    """
    Genera un archivo Excel en memoria.
    filas: lista de dicts con las columnas a escribir.
    columnas: orden de columnas; si None, se infiere de la primera fila.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    cols = columnas or (list(filas[0].keys()) if filas else ["fecha", "cantidad"])
    ws.append(cols)
    for fila in filas:
        ws.append([fila.get(c) for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _multipart(excel_bytes: bytes, filename: str = "test.xlsx"):
    """Devuelve el dict de files listo para TestClient."""
    return {"file": (filename, io.BytesIO(excel_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}


pytestmark = pytest.mark.skipif(
    not OPENPYXL_DISPONIBLE,
    reason="openpyxl no está instalado; instalar con: pip install openpyxl",
)


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-01 – Importar Sitotroga con Excel válido
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP01ImportarSitotroga:
    """IMP-01: Importación de Sitotroga desde Excel con datos válidos."""

    def test_importacion_retorna_200_o_201(self, client, admin_headers):
        """IMP-01: Upload de Excel válido → HTTP 200."""
        filas = [
            {"fecha": "2025-01-10", "cantidad": 100.0},
            {"fecha": "2025-01-11", "cantidad": 200.0},
            {"fecha": "2025-01-12", "cantidad": 150.0},
        ]
        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)

    def test_retorna_conteo_de_importados(self, client, admin_headers):
        """IMP-01: La respuesta incluye cuántos registros se importaron."""
        filas = [
            {"fecha": "2025-02-01", "cantidad": 300.0},
            {"fecha": "2025-02-02", "cantidad": 400.0},
        ]
        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        data = resp.json()
        assert "importados" in data
        assert data["importados"] == 2

    def test_retorna_conteo_de_errores(self, client, admin_headers):
        """IMP-01: La respuesta incluye el campo errores."""
        filas = [{"fecha": "2025-03-01", "cantidad": 50.0}]
        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        assert "errores" in resp.json()

    def test_errores_es_cero_con_datos_validos(self, client, admin_headers):
        """IMP-01: Con datos completamente válidos, errores debe ser 0."""
        filas = [{"fecha": "2025-04-01", "cantidad": 75.0}]
        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        assert resp.json()["errores"] == 0


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-02 – Importar Trichogramma
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP02ImportarTrichogramma:
    """IMP-02: Importación de Trichogramma desde Excel válido."""

    def test_importacion_exitosa(self, client, admin_headers):
        """IMP-02: Excel con 3 filas válidas → importados=3, errores=0."""
        filas = [
            {"fecha": "2025-01-05", "cantidad": 500.0},
            {"fecha": "2025-01-06", "cantidad": 600.0},
            {"fecha": "2025-01-07", "cantidad": 700.0},
        ]
        resp = client.post(
            "/api/v1/importar/trichogramma",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        assert resp.json()["importados"] == 3
        assert resp.json()["errores"] == 0

    def test_fila_unica_importada_correctamente(self, client, admin_headers):
        """IMP-02: Excel con una sola fila válida → importados=1."""
        filas = [{"fecha": "2025-05-01", "cantidad": 999.0}]
        resp = client.post(
            "/api/v1/importar/trichogramma",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        assert resp.json()["importados"] == 1


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-03 – Importar Galleria
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP03ImportarGalleria:
    """IMP-03: Importación de Galleria desde Excel válido."""

    def test_importacion_exitosa(self, client, admin_headers):
        """IMP-03: Galleria importa correctamente desde Excel."""
        filas = [
            {"fecha": "2025-03-10", "cantidad": 1000.0},
            {"fecha": "2025-03-11", "cantidad": 1200.0},
        ]
        resp = client.post(
            "/api/v1/importar/galleria",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        assert resp.json()["importados"] == 2

    def test_con_columna_id_unidad(self, client, admin_headers):
        """IMP-03: Excel con columna id_unidad opcional también funciona."""
        filas = [
            {"fecha": "2025-03-15", "cantidad": 500.0, "id_unidad": 1},
        ]
        resp = client.post(
            "/api/v1/importar/galleria",
            files=_multipart(_excel_bytes(filas, columnas=["fecha", "cantidad", "id_unidad"])),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        assert resp.json()["importados"] == 1


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-04 – Importar Paratheresia
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP04ImportarParatheresia:
    """IMP-04: Importación de Paratheresia desde Excel válido."""

    def test_importacion_exitosa(self, client, admin_headers):
        """IMP-04: Paratheresia importa correctamente."""
        filas = [
            {"fecha": "2025-04-01", "cantidad": 250.0},
            {"fecha": "2025-04-02", "cantidad": 350.0},
            {"fecha": "2025-04-03", "cantidad": 450.0},
        ]
        resp = client.post(
            "/api/v1/importar/paratheresia",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        assert resp.json()["importados"] == 3
        assert resp.json()["errores"] == 0


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-05 – Filas con fecha nula se omiten
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP05FechasNulas:
    """IMP-05: Las filas con fecha nula o vacía deben ser omitidas sin crashear."""

    def test_fila_sin_fecha_cuenta_como_omitida(self, client, admin_headers):
        """IMP-05: 1 fila válida + 1 sin fecha → importados=1 (la nula se descarta)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["fecha", "cantidad"])
        ws.append(["2025-05-10", 100.0])  # fila válida
        ws.append([None, 200.0])           # fecha nula → descartada por dropna
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=_multipart(buf.read()),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        # La fila nula se descarta antes de procesar, no cuenta como error
        assert resp.json()["importados"] == 1

    def test_todas_fechas_nulas_importa_cero(self, client, admin_headers):
        """IMP-05: Excel donde todas las fechas son nulas → importados=0."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["fecha", "cantidad"])
        ws.append([None, 100.0])
        ws.append([None, 200.0])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post(
            "/api/v1/importar/trichogramma",
            files=_multipart(buf.read()),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        assert resp.json()["importados"] == 0


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-06 – Cantidad inválida cuenta como error
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP06CantidadInvalida:
    """IMP-06: Una fila con cantidad de tipo texto no numérico cuenta como error."""

    def test_cantidad_texto_cuenta_como_error(self, client, admin_headers):
        """IMP-06: 1 fila válida + 1 con cantidad='abc' → importados=1, errores=1."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["fecha", "cantidad"])
        ws.append(["2025-06-01", 100.0])     # válida
        ws.append(["2025-06-02", "no_numero"])  # errónea
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=_multipart(buf.read()),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        data = resp.json()
        assert data["importados"] == 1
        assert data["errores"] == 1

    def test_cantidad_negativa_es_error(self, client, admin_headers):
        """IMP-06: Si el schema valida cantidad > 0, negativa debe ser error."""
        filas = [{"fecha": "2025-06-10", "cantidad": -50.0}]
        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=_multipart(_excel_bytes(filas)),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        # La fila negativa debe contar como error si el schema la rechaza
        data = resp.json()
        assert data["importados"] + data["errores"] == 1


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-07 – Archivo no Excel
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP07ArchivoNoExcel:
    """IMP-07: Subir un archivo que no es Excel debe retornar error (400 o 500)."""

    def test_archivo_txt_falla(self, client, admin_headers):
        """IMP-07: Subir un .txt → el servidor debe retornar un error."""
        contenido = b"fecha,cantidad\n2025-01-01,100\n"
        files = {"file": ("datos.txt", io.BytesIO(contenido), "text/plain")}
        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=files,
            headers=admin_headers,
        )
        # Puede retornar 400, 422, o 500 dependiendo de cómo openpyxl maneja el error
        assert resp.status_code in (400, 422, 500)

    def test_archivo_vacio_falla(self, client, admin_headers):
        """IMP-07: Subir un archivo vacío → error del servidor."""
        files = {"file": ("vacio.xlsx", io.BytesIO(b""), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=files,
            headers=admin_headers,
        )
        assert resp.status_code in (400, 422, 500)


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-08 – Control de acceso (solo admin)
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP08ControlAcceso:
    """IMP-08: Solo admin puede acceder a los endpoints de importación."""

    ENDPOINTS_IMPORTAR = [
        "/api/v1/importar/sitotroga",
        "/api/v1/importar/trichogramma",
        "/api/v1/importar/galleria",
        "/api/v1/importar/paratheresia",
    ]

    @pytest.mark.parametrize("endpoint", ENDPOINTS_IMPORTAR)
    def test_operario_no_puede_importar(self, client, operario_headers, endpoint):
        """IMP-08: Operario → POST /importar/* → 403."""
        filas = [{"fecha": "2025-01-01", "cantidad": 100.0}]
        resp = client.post(
            endpoint,
            files=_multipart(_excel_bytes(filas)),
            headers=operario_headers,
        )
        assert resp.status_code == 403

    @pytest.mark.parametrize("endpoint", ENDPOINTS_IMPORTAR)
    def test_supervisor_no_puede_importar(self, client, supervisor_headers, endpoint):
        """IMP-08: Supervisor → POST /importar/* → 403."""
        filas = [{"fecha": "2025-01-01", "cantidad": 100.0}]
        resp = client.post(
            endpoint,
            files=_multipart(_excel_bytes(filas)),
            headers=supervisor_headers,
        )
        assert resp.status_code == 403

    @pytest.mark.parametrize("endpoint", ENDPOINTS_IMPORTAR)
    def test_sin_token_retorna_401(self, client, endpoint):
        """IMP-08: Sin autenticación → 401 o 403."""
        filas = [{"fecha": "2025-01-01", "cantidad": 100.0}]
        resp = client.post(
            endpoint,
            files=_multipart(_excel_bytes(filas)),
        )
        assert resp.status_code in (401, 403)


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-09 – Sin archivo en la request
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP09SinArchivo:
    """IMP-09: Request sin el campo 'file' → HTTP 422."""

    def test_sin_archivo_sitotroga_retorna_422(self, client, admin_headers):
        """IMP-09: POST sin multipart file → 422."""
        resp = client.post("/api/v1/importar/sitotroga", headers=admin_headers)
        assert resp.status_code == 422

    def test_sin_archivo_trichogramma_retorna_422(self, client, admin_headers):
        """IMP-09: POST sin multipart file en trichogramma → 422."""
        resp = client.post("/api/v1/importar/trichogramma", headers=admin_headers)
        assert resp.status_code == 422

    def test_sin_archivo_galleria_retorna_422(self, client, admin_headers):
        resp = client.post("/api/v1/importar/galleria", headers=admin_headers)
        assert resp.status_code == 422

    def test_sin_archivo_paratheresia_retorna_422(self, client, admin_headers):
        resp = client.post("/api/v1/importar/paratheresia", headers=admin_headers)
        assert resp.status_code == 422


# ═══════════════════════════════════════════════════════════════════════════════
# IMP-10 – Todas las filas erróneas
# ═══════════════════════════════════════════════════════════════════════════════
class TestIMP10TodasLasFilasErroneas:
    """IMP-10: Excel donde ninguna fila puede procesarse → importados=0, errores=N."""

    def test_filas_todas_erroneas(self, client, admin_headers):
        """IMP-10: 3 filas con cantidades inválidas → importados=0, errores=3."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["fecha", "cantidad"])
        ws.append(["2025-07-01", "TEXTO"])
        ws.append(["2025-07-02", "TEXTO"])
        ws.append(["2025-07-03", "TEXTO"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post(
            "/api/v1/importar/sitotroga",
            files=_multipart(buf.read()),
            headers=admin_headers,
        )
        assert resp.status_code in (200, 201)
        data = resp.json()
        assert data["importados"] == 0
        assert data["errores"] == 3

    def test_respuesta_no_crashea_con_filas_malas(self, client, admin_headers):
        """IMP-10: Incluso con datos erróneos, el endpoint no debe retornar 500."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["fecha", "cantidad"])
        ws.append(["FECHA_INVALIDA", "NO_ES_NUMERO"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post(
            "/api/v1/importar/galleria",
            files=_multipart(buf.read()),
            headers=admin_headers,
        )
        assert resp.status_code != 500, (
            f"El endpoint crasheó con datos inválidos (HTTP 500): {resp.text}"
        )

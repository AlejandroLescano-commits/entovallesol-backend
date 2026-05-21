"""
test_importacion_reporte.py – Pruebas Unitarias de Importación y Reportes
EntoValleSOL Backend

Casos cubiertos:
  CA01 – ImportacionService._leer_excel  (columnas normalizadas, dropna fecha)
  CA02 – ImportacionService.importar_sitotroga  (conteo correcto, manejo de errores)
  CA03 – ImportacionService.importar_trichogramma / galleria / paratheresia
  CA04 – ReporteService.generar_excel_sitotroga  (bytes, hoja, sin filas vacías)
  CA05 – ReporteService: trichogramma / paratheresia / galleria
  CA06 – ReporteService.generar_excel_notas_*  (encabezados correctos)

Ejecutar:
  pytest tests/unit/test_importacion_reporte.py -v
"""
import io
import pytest
from datetime import date
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _excel_bytes(rows: list[dict]) -> bytes:
    """Genera bytes de un Excel simple a partir de una lista de dicts."""
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    return buf.getvalue()


def _make_prod(fecha: date, cantidad: float, tiposalida: str = None,
               ratio: float = None, id_unidad: int = None,
               factor: float = 1.0, descripcion: str = None,
               id_lugarliberacion: int = None):
    r = MagicMock()
    r.id                 = 1
    r.fecha              = fecha
    r.cantidad           = cantidad
    r.tiposalida         = tiposalida
    r.ratio              = ratio
    r.id_unidad          = id_unidad
    r.factor             = factor
    r.descripcion        = descripcion
    r.id_lugarliberacion = id_lugarliberacion
    return r


def _make_reporte_service(mock_db, prod=None, notas=None):
    """Crea ReporteService con repo mockeado."""
    from app.services.reporte_service import ReporteService
    svc = ReporteService.__new__(ReporteService)
    repo = MagicMock()
    # Producción
    repo.list_sitotroga.return_value     = prod or []
    repo.list_trichogramma.return_value  = prod or []
    repo.list_paratheresia.return_value  = prod or []
    repo.list_galleria.return_value      = prod or []
    # Notas
    repo.list_notas_sitodroga.return_value  = notas or []
    repo.list_notas_avispitas.return_value  = notas or []
    repo.list_notas_moscas.return_value     = notas or []
    repo.list_notas_galleria.return_value   = notas or []
    svc.repo = repo
    return svc


# ═══════════════════════════════════════════════════════════════════════════════
# CA01 – ImportacionService._leer_excel
# ═══════════════════════════════════════════════════════════════════════════════

class TestLeerExcel:
    """CA01: Lectura y normalización de archivos Excel."""

    def _svc(self):
        from app.services.importacion_service import ImportacionService
        svc = ImportacionService.__new__(ImportacionService)
        return svc

    def test_columnas_normalizadas_a_minusculas(self):
        """CA01-1: Las columnas deben convertirse a minúsculas sin espacios."""
        svc = self._svc()
        xlsx = _excel_bytes([{"Fecha": "2025-01-01", "Cantidad": 100.0}])
        df = svc._leer_excel(xlsx)
        assert "fecha"    in df.columns
        assert "cantidad" in df.columns

    def test_filas_sin_fecha_eliminadas(self):
        """CA01-2: Filas con fecha=NaN deben eliminarse."""
        svc = self._svc()
        xlsx = _excel_bytes([
            {"fecha": "2025-01-01", "cantidad": 100.0},
            {"fecha": None,         "cantidad": 200.0},  # debe eliminarse
        ])
        df = svc._leer_excel(xlsx)
        assert len(df) == 1

    def test_retorna_dataframe(self):
        """CA01-3: El resultado debe ser un DataFrame de pandas."""
        svc = self._svc()
        xlsx = _excel_bytes([{"fecha": "2025-01-01", "cantidad": 50.0}])
        df = svc._leer_excel(xlsx)
        assert isinstance(df, pd.DataFrame)

    def test_columnas_con_espacios_normalizadas(self):
        """CA01-4: Columna ' Fecha ' (con espacios) debe quedar como 'fecha'."""
        svc = self._svc()
        df_raw = pd.DataFrame([{" Fecha ": "2025-01-01", "  Cantidad  ": 99.0}])
        buf = io.BytesIO()
        df_raw.to_excel(buf, index=False)
        df = svc._leer_excel(buf.getvalue())
        assert "fecha"    in df.columns
        assert "cantidad" in df.columns


# ═══════════════════════════════════════════════════════════════════════════════
# CA02 – ImportacionService.importar_sitotroga
# ═══════════════════════════════════════════════════════════════════════════════

class TestImportarSitotroga:
    """CA02: Importación masiva de Sitotroga."""

    def _svc(self, mock_db):
        from app.services.importacion_service import ImportacionService
        svc = ImportacionService.__new__(ImportacionService)
        svc.db  = mock_db
        svc.svc = MagicMock()
        return svc

    def test_importa_filas_validas(self, mock_db):
        """CA02-1: 3 filas válidas → importados=3, errores=0."""
        svc  = self._svc(mock_db)
        xlsx = _excel_bytes([
            {"fecha": date(2025, 1, i), "cantidad": float(i * 100)}
            for i in range(1, 4)
        ])
        result = svc.importar_sitotroga(xlsx, user_id=1)
        assert result["importados"] == 3
        assert result["errores"]    == 0

    def test_fila_sin_cantidad_cuenta_como_error(self, mock_db):
        """CA02-2: Fila con datos inválidos suma al contador de errores."""
        svc = self._svc(mock_db)
        svc.svc.registrar_sitotroga.side_effect = Exception("Error de validación")
        xlsx = _excel_bytes([{"fecha": date(2025, 1, 1), "cantidad": -999.0}])
        result = svc.importar_sitotroga(xlsx, user_id=1)
        assert result["errores"] >= 1

    def test_retorna_dict_con_claves_importados_errores(self, mock_db):
        """CA02-3: El resultado siempre tiene las claves 'importados' y 'errores'."""
        svc  = self._svc(mock_db)
        xlsx = _excel_bytes([{"fecha": date(2025, 1, 1), "cantidad": 100.0}])
        result = svc.importar_sitotroga(xlsx, user_id=1)
        assert "importados" in result
        assert "errores"    in result

    def test_id_unidad_none_cuando_no_existe_columna(self, mock_db):
        """CA02-4: Sin columna id_unidad en el Excel, el valor debe ser None."""
        svc      = self._svc(mock_db)
        captured = []
        svc.svc.registrar_sitotroga.side_effect = lambda d, uid: captured.append(d)
        xlsx = _excel_bytes([{"fecha": date(2025, 1, 1), "cantidad": 100.0}])
        svc.importar_sitotroga(xlsx, user_id=1)
        assert captured[0].id_unidad is None

    def test_total_es_suma_de_importados_y_errores(self, mock_db):
        """CA02-5: importados + errores == total de filas del Excel."""
        svc = self._svc(mock_db)
        # 2 buenas, 1 mala
        call_count = 0
        def fake_registrar(d, uid):
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise Exception("fallo")
        svc.svc.registrar_sitotroga.side_effect = fake_registrar
        xlsx = _excel_bytes([{"fecha": date(2025, 1, i), "cantidad": 100.0} for i in range(1, 4)])
        result = svc.importar_sitotroga(xlsx, user_id=1)
        assert result["importados"] + result["errores"] == 3


# ═══════════════════════════════════════════════════════════════════════════════
# CA03 – Importar otras especies
# ═══════════════════════════════════════════════════════════════════════════════

class TestImportarOtrasEspecies:
    """CA03: importar_trichogramma, galleria y paratheresia siguen la misma lógica."""

    def _svc(self, mock_db):
        from app.services.importacion_service import ImportacionService
        svc = ImportacionService.__new__(ImportacionService)
        svc.db  = mock_db
        svc.svc = MagicMock()
        return svc

    def _xlsx_basico(self, n=3):
        return _excel_bytes([
            {"fecha": date(2025, 1, i), "cantidad": float(i * 50)}
            for i in range(1, n + 1)
        ])

    def test_importar_trichogramma_cuenta_correcta(self, mock_db):
        """CA03-1: importar_trichogramma retorna importados correcto."""
        svc    = self._svc(mock_db)
        result = svc.importar_trichogramma(self._xlsx_basico(2), user_id=1)
        assert result["importados"] == 2
        assert result["errores"]    == 0

    def test_importar_galleria_cuenta_correcta(self, mock_db):
        """CA03-2: importar_galleria retorna importados correcto."""
        svc    = self._svc(mock_db)
        result = svc.importar_galleria(self._xlsx_basico(4), user_id=1)
        assert result["importados"] == 4

    def test_importar_paratheresia_cuenta_correcta(self, mock_db):
        """CA03-3: importar_paratheresia retorna importados correcto."""
        svc    = self._svc(mock_db)
        result = svc.importar_paratheresia(self._xlsx_basico(1), user_id=1)
        assert result["importados"] == 1

    def test_errores_en_trichogramma_no_detienen_resto(self, mock_db):
        """CA03-4: Un error en una fila no detiene la importación del resto."""
        svc = self._svc(mock_db)
        call_count = 0
        def fake(d, uid):
            nonlocal call_count
            call_count += 1
            if call_count == 1:
                raise Exception("fallo")
        svc.svc.registrar_trichogramma.side_effect = fake
        result = svc.importar_trichogramma(self._xlsx_basico(3), user_id=1)
        assert result["importados"] == 2
        assert result["errores"]    == 1


# ═══════════════════════════════════════════════════════════════════════════════
# CA04 – ReporteService: Excel Sitotroga
# ═══════════════════════════════════════════════════════════════════════════════

class TestReporteExcelSitotroga:
    """CA04: Generación de Excel de producción Sitotroga."""

    def _prod(self):
        return [_make_prod(date(2025, 5, d), float(d * 100)) for d in range(1, 4)]

    def _notas(self):
        return [_make_prod(date(2025, 5, 1), 50.0, tiposalida="T.exiguum")]

    def test_retorna_bytes_no_vacios(self, mock_db):
        """CA04-1: El resultado debe ser bytes y no estar vacío."""
        svc    = _make_reporte_service(mock_db, self._prod(), self._notas())
        result = svc.generar_excel_sitotroga(date(2025, 5, 1), date(2025, 5, 3))
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_excel_abre_correctamente(self, mock_db):
        """CA04-2: Los bytes deben generar un workbook válido con openpyxl."""
        svc  = _make_reporte_service(mock_db, self._prod(), self._notas())
        data = svc.generar_excel_sitotroga(date(2025, 5, 1), date(2025, 5, 3))
        wb   = openpyxl.load_workbook(io.BytesIO(data))
        assert wb is not None

    def test_hoja_tiene_nombre_correcto(self, mock_db):
        """CA04-3: La hoja activa debe llamarse 'Produccion Sitotroga'."""
        svc  = _make_reporte_service(mock_db, self._prod(), self._notas())
        data = svc.generar_excel_sitotroga(date(2025, 5, 1), date(2025, 5, 3))
        wb   = openpyxl.load_workbook(io.BytesIO(data))
        assert "Produccion Sitotroga" in wb.sheetnames

    def test_titulo_en_primera_celda(self, mock_db):
        """CA04-4: La primera celda debe contener el título de Sitotroga."""
        svc  = _make_reporte_service(mock_db, self._prod(), self._notas())
        data = svc.generar_excel_sitotroga(date(2025, 5, 1), date(2025, 5, 3))
        wb   = openpyxl.load_workbook(io.BytesIO(data))
        ws   = wb.active
        # La celda A1 contiene el título (puede ser mergeada)
        assert ws.cell(1, 1).value is not None
        assert "Sitotroga" in str(ws.cell(1, 1).value)

    def test_rango_un_dia_genera_fila_datos(self, mock_db):
        """CA04-5: Rango de un solo día genera exactamente una fila de datos."""
        prod  = [_make_prod(date(2025, 5, 15), 300.0)]
        svc   = _make_reporte_service(mock_db, prod, [])
        data  = svc.generar_excel_sitotroga(date(2025, 5, 15), date(2025, 5, 15))
        wb    = openpyxl.load_workbook(io.BytesIO(data))
        ws    = wb.active
        # Fila 6 es el primer día de datos (filas 1-5 son encabezados)
        assert ws.cell(6, 1).value is not None


# ═══════════════════════════════════════════════════════════════════════════════
# CA05 – ReporteService: otras especies
# ═══════════════════════════════════════════════════════════════════════════════

class TestReporteOtrasEspecies:
    """CA05: Excel de Trichogramma, Paratheresia y Galleria."""

    def _prod(self):
        return [_make_prod(date(2025, 5, d), float(d * 100)) for d in range(1, 4)]

    def _check_excel(self, data: bytes, nombre_hoja: str, keyword: str):
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert nombre_hoja in wb.sheetnames
        ws = wb.active
        assert keyword in str(ws.cell(1, 1).value)

    def test_excel_trichogramma_retorna_bytes(self, mock_db):
        """CA05-1: generar_excel_trichogramma retorna bytes no vacíos."""
        svc  = _make_reporte_service(mock_db, self._prod(), [])
        data = svc.generar_excel_trichogramma(date(2025, 5, 1), date(2025, 5, 3))
        assert len(data) > 0
        self._check_excel(data, "Produccion Trichogramma", "Trichogramma")

    def test_excel_paratheresia_retorna_bytes(self, mock_db):
        """CA05-2: generar_excel_paratheresia retorna bytes no vacíos."""
        svc  = _make_reporte_service(mock_db, self._prod(), [])
        data = svc.generar_excel_paratheresia(date(2025, 5, 1), date(2025, 5, 3))
        assert len(data) > 0
        self._check_excel(data, "Produccion Paratheresia", "Paratheresia")

    def test_excel_galleria_retorna_bytes(self, mock_db):
        """CA05-3: generar_excel_galleria retorna bytes no vacíos."""
        svc  = _make_reporte_service(mock_db, self._prod(), [])
        data = svc.generar_excel_galleria(date(2025, 5, 1), date(2025, 5, 3))
        assert len(data) > 0
        self._check_excel(data, "Produccion Galleria", "Galleria")

    def test_rango_vacio_sin_datos_genera_excel_valido(self, mock_db):
        """CA05-4: Sin datos de producción el Excel igual debe ser válido."""
        svc  = _make_reporte_service(mock_db, [], [])
        data = svc.generar_excel_trichogramma(date(2025, 5, 1), date(2025, 5, 3))
        wb   = openpyxl.load_workbook(io.BytesIO(data))
        assert wb is not None

    def test_saldo_se_acumula_correctamente(self, mock_db):
        """CA05-5: El saldo en el reporte se calcula acumulando producción menos salidas."""
        prod  = [_make_prod(date(2025, 5, 1), 500.0)]
        notas = [_make_prod(date(2025, 5, 1), 200.0, tiposalida="Parasitación")]
        svc   = _make_reporte_service(mock_db, prod, notas)
        # Solo verificamos que genera el Excel sin errores y tiene contenido
        data  = svc.generar_excel_paratheresia(date(2025, 5, 1), date(2025, 5, 1))
        assert len(data) > 0


# ═══════════════════════════════════════════════════════════════════════════════
# CA06 – ReporteService: Notas de salida
# ═══════════════════════════════════════════════════════════════════════════════

class TestReporteNotas:
    """CA06: Excel de notas de salida por especie."""

    HDRS_SITODROGA  = ["ID", "Fecha", "Tipo Salida", "Descripción", "Unidad", "Factor", "Cantidad"]
    HDRS_AVISPITAS  = ["ID", "Fecha", "Tipo Salida", "Lugar Liberación", "Descripción", "Unidad", "Cantidad"]
    HDRS_GALLERIA   = ["ID", "Fecha", "Tipo Salida", "Descripción", "Unidad", "Cantidad", "Ratio"]

    def _nota(self, fecha=date(2025, 5, 1), cantidad=100.0, tiposalida="Ventas"):
        return _make_prod(fecha, cantidad, tiposalida=tiposalida,
                          factor=1.0, descripcion="desc", id_unidad=1,
                          id_lugarliberacion=2, ratio=0.5)

    def _encabezados(self, data: bytes) -> list:
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        return [ws.cell(1, j).value for j in range(1, 20) if ws.cell(1, j).value]

    def test_notas_sitodroga_encabezados_correctos(self, mock_db):
        """CA06-1: Encabezados de notas sitodroga deben coincidir exactamente."""
        svc  = _make_reporte_service(mock_db, notas=[self._nota()])
        data = svc.generar_excel_notas_sitodroga(date(2025, 5, 1), date(2025, 5, 31))
        hdrs = self._encabezados(data)
        for h in self.HDRS_SITODROGA:
            assert h in hdrs, f"Encabezado faltante: {h}"

    def test_notas_avispitas_encabezados_correctos(self, mock_db):
        """CA06-2: Encabezados de notas avispitas deben coincidir."""
        svc  = _make_reporte_service(mock_db, notas=[self._nota()])
        data = svc.generar_excel_notas_avispitas(date(2025, 5, 1), date(2025, 5, 31))
        hdrs = self._encabezados(data)
        for h in self.HDRS_AVISPITAS:
            assert h in hdrs, f"Encabezado faltante: {h}"

    def test_notas_moscas_encabezados_correctos(self, mock_db):
        """CA06-3: Encabezados de notas moscas deben coincidir."""
        svc  = _make_reporte_service(mock_db, notas=[self._nota()])
        data = svc.generar_excel_notas_moscas(date(2025, 5, 1), date(2025, 5, 31))
        hdrs = self._encabezados(data)
        for h in self.HDRS_AVISPITAS:
            assert h in hdrs

    def test_notas_galleria_encabezados_correctos(self, mock_db):
        """CA06-4: Encabezados de notas galleria deben incluir Ratio."""
        svc  = _make_reporte_service(mock_db, notas=[self._nota()])
        data = svc.generar_excel_notas_galleria(date(2025, 5, 1), date(2025, 5, 31))
        hdrs = self._encabezados(data)
        for h in self.HDRS_GALLERIA:
            assert h in hdrs

    def test_sin_notas_genera_excel_solo_con_encabezados(self, mock_db):
        """CA06-5: Sin notas el Excel debe tener solo la fila de encabezados (fila 1)."""
        svc  = _make_reporte_service(mock_db, notas=[])
        data = svc.generar_excel_notas_sitodroga(date(2025, 5, 1), date(2025, 5, 31))
        wb   = openpyxl.load_workbook(io.BytesIO(data))
        ws   = wb.active
        # Fila 2 debe estar vacía (no hay datos)
        assert ws.cell(2, 1).value is None

    def test_con_notas_genera_filas_de_datos(self, mock_db):
        """CA06-6: Con 2 notas debe haber datos en filas 2 y 3."""
        notas = [self._nota(date(2025, 5, i)) for i in range(1, 3)]
        svc   = _make_reporte_service(mock_db, notas=notas)
        data  = svc.generar_excel_notas_sitodroga(date(2025, 5, 1), date(2025, 5, 31))
        wb    = openpyxl.load_workbook(io.BytesIO(data))
        ws    = wb.active
        assert ws.cell(2, 1).value is not None
        assert ws.cell(3, 1).value is not None

from sqlalchemy.orm import Session
from datetime import date, timedelta
from calendar import monthrange
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from app.infrastructure.repositories.produccion_repository import ProduccionRepository

MESES_ES = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# ── Estilos ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
TOTAL_BG   = "D6E4F0"
WHITE      = "FFFFFF"

thin = Side(style="thin")
BORDER_ALL  = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_BOLD = Border(left=Side(style="medium"), right=Side(style="medium"),
                     top=Side(style="medium"), bottom=Side(style="medium"))

def _font(bold=False, color="000000", size=9, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _style(ws, row, col, value=None, bold=False, fg=None, font_color="000000",
           h="center", wrap=False, size=9, italic=False, border=True, num_fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = _font(bold=bold, color=font_color, size=size, italic=italic)
    if fg:
        c.fill = _fill(fg)
    c.alignment = _align(h=h, wrap=wrap)
    if border:
        c.border = BORDER_ALL
    if num_fmt:
        c.number_format = num_fmt
    return c

def _title_row(ws, row, text, n_cols, size=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, color=WHITE, size=size)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align(wrap=False)
    ws.row_dimensions[row].height = 16

def _col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _all_days(fi: date, ff: date):
    d, days = fi, []
    while d <= ff:
        days.append(d)
        d += timedelta(days=1)
    return days

def _to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# SITOTROGA  ─  Producción Huevos (gramos)
# Salida cols: T.exiguum | T.pretiosum | Crysopas | Infestación | Ventas
# ═══════════════════════════════════════════════════════════════════════════════
SITO_SALIDAS = ["T.exiguum", "T.pretiosum", "Crysopas", "Infestación", "Ventas"]

def _build_sitotroga(ws, fi: date, ff: date, prod_rows, nota_rows):
    mes_label = MESES_ES[fi.month]
    n_cols = 3 + len(SITO_SALIDAS) + 2  # Fecha+Unidad+Prod | salidas | Total+Saldo

    _title_row(ws, 1, "PRODUCCION DE HUEVOS DE Sitotroga cerealella", n_cols, size=11)
    _title_row(ws, 2, "UNIDAD = GRAMOS", n_cols, size=10)
    _title_row(ws, 3, mes_label.upper(), n_cols, size=10)

    # Encabezado col-4 merge "Salida"
    sc = 4  # start col for salidas
    ec = sc + len(SITO_SALIDAS) - 1
    ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)
    for col in [1,2,3]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    ws.merge_cells(start_row=4, start_column=ec+1, end_row=5, end_column=ec+1)
    ws.merge_cells(start_row=4, start_column=ec+2, end_row=5, end_column=ec+2)

    hdrs4 = ["Fecha de ingreso", "Unidad", "Producción día", "Salida"]
    for i, h in enumerate(["Fecha de ingreso","Unidad","Producción día"], 1):
        _style(ws, 4, i, h, bold=True, fg=LIGHT_BLUE, wrap=True)
    _style(ws, 4, sc, "Salida", bold=True, fg=LIGHT_BLUE)
    _style(ws, 4, ec+1, "Total", bold=True, fg=LIGHT_BLUE)
    _style(ws, 4, ec+2, "Saldo", bold=True, fg=LIGHT_BLUE)

    for i, s in enumerate(SITO_SALIDAS, sc):
        _style(ws, 5, i, s, bold=True, fg=LIGHT_BLUE, wrap=True)

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 28

    # Índices de datos
    prod_by_day = defaultdict(float)
    for r in prod_rows:
        prod_by_day[r.fecha] += r.cantidad

    notas_by_day = defaultdict(lambda: defaultdict(float))
    for r in nota_rows:
        notas_by_day[r.fecha][r.tiposalida] += r.cantidad

    days = _all_days(fi, ff)
    saldo = 0.0
    data_start = 6
    total_prod = 0.0
    total_salidas = defaultdict(float)
    total_total = 0.0

    for i, d in enumerate(days):
        row = data_start + i
        prod = prod_by_day.get(d, None)
        salidas = {s: notas_by_day[d].get(s, None) for s in SITO_SALIDAS}
        total_sal = sum(v for v in salidas.values() if v)
        saldo = saldo + (prod or 0) - total_sal

        _style(ws, row, 1, f"{d.day}-{d.strftime('%b')}", h="center")
        _style(ws, row, 2, "gramos")
        _style(ws, row, 3, prod if prod else None)
        for j, s in enumerate(SITO_SALIDAS, sc):
            _style(ws, row, j, salidas[s] if salidas[s] else None)
        _style(ws, row, ec+1, total_sal if total_sal else None)
        _style(ws, row, ec+2, round(saldo, 2))

        if prod: total_prod += prod
        for s in SITO_SALIDAS:
            if salidas[s]: total_salidas[s] += salidas[s]
        total_total += total_sal

    # Fila totales
    tr = data_start + len(days)
    _style(ws, tr, 1, "TOTAL", bold=True, fg=TOTAL_BG)
    _style(ws, tr, 2, "", fg=TOTAL_BG)
    _style(ws, tr, 3, round(total_prod, 2), bold=True, fg=TOTAL_BG)
    for j, s in enumerate(SITO_SALIDAS, sc):
        v = total_salidas[s]
        _style(ws, tr, j, round(v, 2) if v else None, bold=True, fg=TOTAL_BG)
    _style(ws, tr, ec+1, round(total_total, 2), bold=True, fg=TOTAL_BG)
    _style(ws, tr, ec+2, "", fg=TOTAL_BG)

    _col_widths(ws, [12, 9, 12] + [12]*len(SITO_SALIDAS) + [10, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# TRICHOGRAMMA  ─  Avispitas (pulg²)
# Salida cols: Parasitación | San Ricardo A | San Ricardo B | Quemazón |
#              Trapiche | Segundo Jirón | Pabellón alto | Sacachique | Encantada | Venta
# ═══════════════════════════════════════════════════════════════════════════════
TRICH_SALIDAS = ["Parasitación","San Ricardo A","San Ricardo B","Quemazón",
                 "Trapiche","Segundo Jirón","Pabellón alto","Sacachique","Encantada","Venta"]

def _build_trichogramma(ws, fi, ff, prod_rows, nota_rows):
    mes_label = MESES_ES[fi.month]
    n_cols = 3 + len(TRICH_SALIDAS) + 2

    _title_row(ws, 1, "PRODUCCION DE AVISPITAS  Trichogramma exiguum", n_cols)
    _title_row(ws, 2, "UNIDAD = pulg2", n_cols, size=10)
    _title_row(ws, 3, mes_label.upper(), n_cols, size=10)

    sc = 4; ec = sc + len(TRICH_SALIDAS) - 1
    for col in [1,2,3]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)
    ws.merge_cells(start_row=4, start_column=ec+1, end_row=5, end_column=ec+1)
    ws.merge_cells(start_row=4, start_column=ec+2, end_row=5, end_column=ec+2)

    for i, h in enumerate(["Fecha de ingreso","Unidad","Producción día – pulg2"], 1):
        _style(ws, 4, i, h, bold=True, fg=LIGHT_BLUE, wrap=True)
    _style(ws, 4, sc, "Salida", bold=True, fg=LIGHT_BLUE)
    _style(ws, 4, ec+1, "Total", bold=True, fg=LIGHT_BLUE)
    _style(ws, 4, ec+2, "Saldo", bold=True, fg=LIGHT_BLUE)
    for j, s in enumerate(TRICH_SALIDAS, sc):
        _style(ws, 5, j, s, bold=True, fg=LIGHT_BLUE, wrap=True)

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 28

    prod_by_day = defaultdict(float)
    for r in prod_rows:
        prod_by_day[r.fecha] += r.cantidad

    notas_by_day = defaultdict(lambda: defaultdict(float))
    for r in nota_rows:
        notas_by_day[r.fecha][r.tiposalida] += r.cantidad

    days = _all_days(fi, ff)
    saldo = 0.0
    total_prod = 0.0
    total_salidas = defaultdict(float)
    total_total = 0.0

    for i, d in enumerate(days):
        row = 6 + i
        prod = prod_by_day.get(d, None)
        salidas = {s: notas_by_day[d].get(s, None) for s in TRICH_SALIDAS}
        total_sal = sum(v for v in salidas.values() if v)
        saldo = saldo + (prod or 0) - total_sal

        _style(ws, row, 1, f"{d.day}-{d.strftime('%b')}", h="center")
        _style(ws, row, 2, "pulg2")
        _style(ws, row, 3, prod if prod else None)
        for j, s in enumerate(TRICH_SALIDAS, sc):
            _style(ws, row, j, salidas[s] if salidas[s] else None)
        _style(ws, row, ec+1, total_sal if total_sal else None)
        _style(ws, row, ec+2, round(saldo, 2))

        if prod: total_prod += prod
        for s in TRICH_SALIDAS:
            if salidas[s]: total_salidas[s] += salidas[s]
        total_total += total_sal

    tr = 6 + len(days)
    _style(ws, tr, 1, "TOTAL", bold=True, fg=TOTAL_BG)
    _style(ws, tr, 2, "", fg=TOTAL_BG)
    _style(ws, tr, 3, round(total_prod, 2), bold=True, fg=TOTAL_BG)
    for j, s in enumerate(TRICH_SALIDAS, sc):
        v = total_salidas[s]
        _style(ws, tr, j, round(v, 2) if v else None, bold=True, fg=TOTAL_BG)
    _style(ws, tr, ec+1, round(total_total, 2), bold=True, fg=TOTAL_BG)
    _style(ws, tr, ec+2, "", fg=TOTAL_BG)

    _col_widths(ws, [12, 9, 14] + [11]*len(TRICH_SALIDAS) + [10, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# PARATHERESIA  ─  Moscas (parejas)
# Salida cols: Parasitación | Quemazón | Segundo Jirón | San Ricardo A |
#              San Ricardo B | Sacachique | Pabellón alto | Trapiche | Venta
# ═══════════════════════════════════════════════════════════════════════════════
PARA_SALIDAS = ["Parasitación","Quemazón","Segundo Jirón","San Ricardo A",
                "San Ricardo B","Sacachique","Pabellón alto","Trapiche","Venta"]

def _build_paratheresia(ws, fi, ff, prod_rows, nota_rows):
    mes_label = MESES_ES[fi.month]
    n_cols = 3 + len(PARA_SALIDAS) + 2

    _title_row(ws, 1, "PRODUCCION DE MOSCAS  Paratheresia claripalpis", n_cols)
    _title_row(ws, 2, "UNIDAD = parejas", n_cols, size=10)
    _title_row(ws, 3, mes_label.upper(), n_cols, size=10)

    sc = 4; ec = sc + len(PARA_SALIDAS) - 1
    for col in [1,2,3]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)
    ws.merge_cells(start_row=4, start_column=ec+1, end_row=5, end_column=ec+1)
    ws.merge_cells(start_row=4, start_column=ec+2, end_row=5, end_column=ec+2)

    for i, h in enumerate(["Fecha de ingreso","Unidad","Producción día – parejas"], 1):
        _style(ws, 4, i, h, bold=True, fg=LIGHT_BLUE, wrap=True)
    _style(ws, 4, sc, "Salida", bold=True, fg=LIGHT_BLUE)
    _style(ws, 4, ec+1, "Total", bold=True, fg=LIGHT_BLUE)
    _style(ws, 4, ec+2, "Saldo", bold=True, fg=LIGHT_BLUE)
    for j, s in enumerate(PARA_SALIDAS, sc):
        _style(ws, 5, j, s, bold=True, fg=LIGHT_BLUE, wrap=True)

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 28

    prod_by_day = defaultdict(float)
    for r in prod_rows:
        prod_by_day[r.fecha] += r.cantidad

    notas_by_day = defaultdict(lambda: defaultdict(float))
    for r in nota_rows:
        notas_by_day[r.fecha][r.tiposalida] += r.cantidad

    days = _all_days(fi, ff)
    saldo = 0.0
    total_prod = 0.0
    total_salidas = defaultdict(float)
    total_total = 0.0

    for i, d in enumerate(days):
        row = 6 + i
        prod = prod_by_day.get(d, None)
        salidas = {s: notas_by_day[d].get(s, None) for s in PARA_SALIDAS}
        total_sal = sum(v for v in salidas.values() if v)
        saldo = saldo + (prod or 0) - total_sal

        _style(ws, row, 1, f"{d.day}-{d.strftime('%b')}", h="center")
        _style(ws, row, 2, "parejas")
        _style(ws, row, 3, prod if prod else None)
        for j, s in enumerate(PARA_SALIDAS, sc):
            _style(ws, row, j, salidas[s] if salidas[s] else None)
        _style(ws, row, ec+1, total_sal if total_sal else None)
        _style(ws, row, ec+2, round(saldo, 2))

        if prod: total_prod += prod
        for s in PARA_SALIDAS:
            if salidas[s]: total_salidas[s] += salidas[s]
        total_total += total_sal

    tr = 6 + len(days)
    _style(ws, tr, 1, "TOTAL", bold=True, fg=TOTAL_BG)
    _style(ws, tr, 2, "", fg=TOTAL_BG)
    _style(ws, tr, 3, round(total_prod, 2), bold=True, fg=TOTAL_BG)
    for j, s in enumerate(PARA_SALIDAS, sc):
        v = total_salidas[s]
        _style(ws, tr, j, round(v, 2) if v else None, bold=True, fg=TOTAL_BG)
    _style(ws, tr, ec+1, round(total_total, 2), bold=True, fg=TOTAL_BG)
    _style(ws, tr, ec+2, "", fg=TOTAL_BG)

    _col_widths(ws, [12, 9, 14] + [11]*len(PARA_SALIDAS) + [10, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# GALLERIA  ─  (Unidades)
# Salida cols: Paratheresia | Instalación | Ventas
# ═══════════════════════════════════════════════════════════════════════════════
GALL_SALIDAS = ["Paratheresia", "Instalación", "Ventas"]

def _build_galleria(ws, fi, ff, prod_rows, nota_rows):
    mes_label = MESES_ES[fi.month]
    n_cols = 3 + len(GALL_SALIDAS) + 2

    _title_row(ws, 1, "PRODUCCION DE Galleria melonella", n_cols)
    _title_row(ws, 2, "UNIDAD = Unidades", n_cols, size=10)
    _title_row(ws, 3, mes_label.upper(), n_cols, size=10)

    sc = 4; ec = sc + len(GALL_SALIDAS) - 1
    for col in [1,2,3]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)
    ws.merge_cells(start_row=4, start_column=ec+1, end_row=5, end_column=ec+1)
    ws.merge_cells(start_row=4, start_column=ec+2, end_row=5, end_column=ec+2)

    for i, h in enumerate(["Fecha de ingreso","Unidad","Producción día"], 1):
        _style(ws, 4, i, h, bold=True, fg=LIGHT_BLUE, wrap=True)
    _style(ws, 4, sc, "Salida", bold=True, fg=LIGHT_BLUE)
    _style(ws, 4, ec+1, "Total", bold=True, fg=LIGHT_BLUE)
    _style(ws, 4, ec+2, "Saldo", bold=True, fg=LIGHT_BLUE)
    for j, s in enumerate(GALL_SALIDAS, sc):
        _style(ws, 5, j, s, bold=True, fg=LIGHT_BLUE, wrap=True)

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 28

    prod_by_day = defaultdict(float)
    for r in prod_rows:
        prod_by_day[r.fecha] += r.cantidad

    notas_by_day = defaultdict(lambda: defaultdict(float))
    for r in nota_rows:
        notas_by_day[r.fecha][r.tiposalida] += r.cantidad

    days = _all_days(fi, ff)
    saldo = 0.0
    total_prod = 0.0
    total_salidas = defaultdict(float)
    total_total = 0.0

    for i, d in enumerate(days):
        row = 6 + i
        prod = prod_by_day.get(d, None)
        salidas = {s: notas_by_day[d].get(s, None) for s in GALL_SALIDAS}
        total_sal = sum(v for v in salidas.values() if v)
        saldo = saldo + (prod or 0) - total_sal

        _style(ws, row, 1, f"{d.day}-{d.strftime('%b')}", h="center")
        _style(ws, row, 2, "Unidad")
        _style(ws, row, 3, prod if prod else None)
        for j, s in enumerate(GALL_SALIDAS, sc):
            _style(ws, row, j, salidas[s] if salidas[s] else None)
        _style(ws, row, ec+1, total_sal if total_sal else None)
        _style(ws, row, ec+2, round(saldo, 2))

        if prod: total_prod += prod
        for s in GALL_SALIDAS:
            if salidas[s]: total_salidas[s] += salidas[s]
        total_total += total_sal

    tr = 6 + len(days)
    _style(ws, tr, 1, "TOTAL", bold=True, fg=TOTAL_BG)
    _style(ws, tr, 2, "", fg=TOTAL_BG)
    _style(ws, tr, 3, round(total_prod, 2), bold=True, fg=TOTAL_BG)
    for j, s in enumerate(GALL_SALIDAS, sc):
        v = total_salidas[s]
        _style(ws, tr, j, round(v, 2) if v else None, bold=True, fg=TOTAL_BG)
    _style(ws, tr, ec+1, round(total_total, 2), bold=True, fg=TOTAL_BG)
    _style(ws, tr, ec+2, "", fg=TOTAL_BG)

    _col_widths(ws, [12, 9, 12, 14, 12, 10, 10, 10])


# ═══════════════════════════════════════════════════════════════════════════════
# SERVICE
# ═══════════════════════════════════════════════════════════════════════════════
class ReporteService:
    def __init__(self, db: Session):
        self.repo = ProduccionRepository(db)

    def _rango_mes(self, mes: int, anio: int):
        fi = date(anio, mes, 1)
        ff = date(anio, mes, monthrange(anio, mes)[1])
        return fi, ff

    # ── Producción mensual ────────────────────────────────────────────────────
    def generar_excel_sitotroga(self, fi: date, ff: date) -> bytes:
        prod  = self.repo.list_sitotroga(fi, ff)
        notas = self.repo.list_notas_sitodroga(fi, ff)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Produccion Sitotroga"
        _build_sitotroga(ws, fi, ff, prod, notas)
        return _to_bytes(wb)

    def generar_excel_trichogramma(self, fi: date, ff: date) -> bytes:
        prod  = self.repo.list_trichogramma(fi, ff)
        notas = self.repo.list_notas_avispitas(fi, ff)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Produccion Trichogramma"
        _build_trichogramma(ws, fi, ff, prod, notas)
        return _to_bytes(wb)

    def generar_excel_paratheresia(self, fi: date, ff: date) -> bytes:
        prod  = self.repo.list_paratheresia(fi, ff)
        notas = self.repo.list_notas_moscas(fi, ff)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Produccion Paratheresia"
        _build_paratheresia(ws, fi, ff, prod, notas)
        return _to_bytes(wb)

    def generar_excel_galleria(self, fi: date, ff: date) -> bytes:
        prod  = self.repo.list_galleria(fi, ff)
        notas = self.repo.list_notas_galleria(fi, ff)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Produccion Galleria"
        _build_galleria(ws, fi, ff, prod, notas)
        return _to_bytes(wb)

    # ── Notas de salida (sin cambios, solo formato mejorado) ──────────────────
    def generar_excel_notas_sitodroga(self, fi: date, ff: date) -> bytes:
        rows = self.repo.list_notas_sitodroga(fi, ff)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Notas Sitodroga"
        hdrs = ["ID","Fecha","Tipo Salida","Descripción","Unidad","Factor","Cantidad"]
        for j, h in enumerate(hdrs, 1):
            _style(ws, 1, j, h, bold=True, fg=LIGHT_BLUE)
        for i, r in enumerate(rows, 2):
            for j, v in enumerate([r.id, str(r.fecha), r.tiposalida, r.descripcion,
                                    r.id_unidad, r.factor, r.cantidad], 1):
                _style(ws, i, j, v)
        return _to_bytes(wb)

    def generar_excel_notas_avispitas(self, fi: date, ff: date) -> bytes:
        rows = self.repo.list_notas_avispitas(fi, ff)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Notas Avispitas"
        hdrs = ["ID","Fecha","Tipo Salida","Lugar Liberación","Descripción","Unidad","Cantidad"]
        for j, h in enumerate(hdrs, 1):
            _style(ws, 1, j, h, bold=True, fg=LIGHT_BLUE)
        for i, r in enumerate(rows, 2):
            for j, v in enumerate([r.id, str(r.fecha), r.tiposalida, r.id_lugarliberacion,
                                    r.descripcion, r.id_unidad, r.cantidad], 1):
                _style(ws, i, j, v)
        return _to_bytes(wb)

    def generar_excel_notas_moscas(self, fi: date, ff: date) -> bytes:
        rows = self.repo.list_notas_moscas(fi, ff)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Notas Moscas"
        hdrs = ["ID","Fecha","Tipo Salida","Lugar Liberación","Descripción","Unidad","Cantidad"]
        for j, h in enumerate(hdrs, 1):
            _style(ws, 1, j, h, bold=True, fg=LIGHT_BLUE)
        for i, r in enumerate(rows, 2):
            for j, v in enumerate([r.id, str(r.fecha), r.tiposalida, r.id_lugarliberacion,
                                    r.descripcion, r.id_unidad, r.cantidad], 1):
                _style(ws, i, j, v)
        return _to_bytes(wb)

    def generar_excel_notas_galleria(self, fi: date, ff: date) -> bytes:
        rows = self.repo.list_notas_galleria(fi, ff)
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = "Notas Galleria"
        hdrs = ["ID","Fecha","Tipo Salida","Descripción","Unidad","Cantidad","Ratio"]
        for j, h in enumerate(hdrs, 1):
            _style(ws, 1, j, h, bold=True, fg=LIGHT_BLUE)
        for i, r in enumerate(rows, 2):
            for j, v in enumerate([r.id, str(r.fecha), r.tiposalida, r.descripcion,
                                    r.id_unidad, r.cantidad, r.ratio], 1):
                _style(ws, i, j, v)
        return _to_bytes(wb)
